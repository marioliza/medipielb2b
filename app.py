import streamlit as st
import pandas as pd
import tempfile
import os
import zipfile
from datetime import datetime
from openpyxl import load_workbook
import shutil

def empaquetar_ordenes_optimo(df_agrupado, max_filas=100):
    bloques = []
    bloques_filas = []
    ordenes = [(orden, grupo) for orden, grupo in df_agrupado.groupby('numero_externo')]
    ordenes.sort(key=lambda x: len(x[1]), reverse=True)

    for orden, grupo in ordenes:
        n_filas = len(grupo)
        colocado = False
        for i in range(len(bloques)):
            if bloques_filas[i] + n_filas <= max_filas:
                bloques[i] = pd.concat([bloques[i], grupo], ignore_index=True)
                bloques_filas[i] += n_filas
                colocado = True
                break
        if not colocado:
            bloques.append(grupo.copy())
            bloques_filas.append(n_filas)

    return [(i+1, bloque) for i, bloque in enumerate(bloques)]

def exportar_bloques_a_template(bloques, plantilla_path, carpeta_salida, ciudad):
    fecha_str = datetime.now().strftime('%Y-%m-%d')
    ciudad_code = ciudad[6:9].upper().replace(" ", "")
    archivos_generados = []

    for archivo_id, bloque in bloques:
        wb = load_workbook(plantilla_path)
        ws = wb['Template']

        column_map = {}
        for col in ws.iter_cols(min_row=2, max_row=2):
            header = str(col[0].value).strip()
            if header:
                column_map[header] = col[0].column_letter

        for i, row in bloque.iterrows():
            fila_excel = i + 3
            ws[f"{column_map['Número de orden externo']}{fila_excel}"] = row['numero_externo']
            ws[f"{column_map['CEDIS de origen']}{fila_excel}"] = row['nombre_bodega']
            ws[f"{column_map['Destinatario']}{fila_excel}"] = row['destinatario']
            ws[f"{column_map['Nombre punto de entrega']}{fila_excel}"] = row['pdv']
            ws[f"{column_map['Alistamiento']}{fila_excel}"] = "Reservar ahora y alistar después"
            ws[f"{column_map['Método de envío']}{fila_excel}"] = "Estándar B2B (Local y Nacional) *"
            ws[f"{column_map['SKU o Código Melonn del producto']}{fila_excel}"] = row['sku']
            ws[f"{column_map['Cantidad']}{fila_excel}"] = row['cantidad']


        nombre_archivo = f"template_{fecha_str}_{ciudad_code}_{str(archivo_id).zfill(3)}.xlsx"
        output_path = os.path.join(carpeta_salida, nombre_archivo)
        wb.save(output_path)
        archivos_generados.append(output_path)

    return archivos_generados

def procesar_archivo_medipiel(archivo_path, plantilla_path):
    xls = pd.ExcelFile(archivo_path)
    bodegas_dict = {
        'ME004': 'Cali #2 - Cámbulos', #--------------------------------------------------------
        'ME002': 'Medellin #2 - Sabaneta Mayorca',
        'ME005': 'Barranquilla #1 - Granadillo',
        'ME003': 'Bogotá #2 - Montevideo'
    }
    hojas = {'Melon Sabaneta', 'Melon Bogotá', 'Melon Cali', 'melon Barranquilla'}
    hojas_normalizadas = {h.lower().strip(): h for h in xls.sheet_names}
    print(hojas_normalizadas)
    hojas_validas = [hojas_normalizadas[h.lower()] for h in hojas if h.lower() in hojas_normalizadas]

    resumen_global = []
    archivos_exportados = []

    with tempfile.TemporaryDirectory() as carpeta_salida:
        for hoja in hojas_validas:
            df = pd.read_excel(xls, sheet_name=hoja)
            print(df.columns)
            col_orden_externa = [c for c in df.columns if 'orden externa' in c.lower()][0]
            col_destinatario = [c for c in df.columns if 'tienda' in c.lower() or 'desc.' in c.lower()][0]
            col_bodega =        [c for c in df.columns if 'bod. salida' in c.lower() or 'salida' in c.lower()][0]
            col_cantidad      = [c for c in df.columns if 'cant' in c.lower()][0]
            col_sku           = [c for c in df.columns if 'codigo' in c.lower()][0]
            col_ceco =          [c for c in df.columns if 'bod. entrada' in c.lower() or 'entrada' in c.lower()][0]

            df['numero_externo'] = df[col_orden_externa].astype(str).str.strip()
            df['destinatario']   = df[col_destinatario].astype(str).str.strip()
            df['bodega']         = df[col_bodega].astype(str).str.strip()
            df['cantidad']       = df[col_cantidad]
            df['sku']            = df[col_sku].astype(str).str.strip()
            df['nombre_bodega']  = df['bodega'].map(bodegas_dict)
            df['entrada'] =        df[col_ceco].astype(str).str.strip()


            #------------
            # === HOMOLOGACIÓN de destinatario =============================================
            homologos_df = pd.read_excel('Homologos_.xlsx')
            homologos_df['Ceco'] = homologos_df['Ceco'].astype(str)
            homologos_df['destinatario_origen_norm'] = homologos_df['Ceco'].str.strip()

            map_dest = dict(zip(
                homologos_df['destinatario_origen_norm'],
                homologos_df['Homologo Melonn / Destinatario']
            ))

            pdv = dict(zip(
                homologos_df['destinatario_origen_norm'],
                homologos_df['PDV']
            ))

            df['destinatario_norm'] = df['entrada'].str.strip().str[2:]
            df['destinatario_homologado'] = df['destinatario_norm'].map(map_dest).fillna(df['destinatario'])
            df['pdv'] = df['destinatario_norm'].map(pdv).fillna(df['destinatario'])

            # ⚠️ VALIDACIÓN OPCIONAL: mostrar los no homologados
            no_homologados = df[df['destinatario_norm'].isin(
                set(df['destinatario_norm']) - set(map_dest.keys())
            )]

            if not no_homologados.empty:
                st.warning(f"⚠️ {no_homologados['destinatario'].nunique()} destinatarios no tienen homologación.")
                st.dataframe(no_homologados[['destinatario']].drop_duplicates())

            # Reasignación final
            df['destinatario'] = df['destinatario_homologado']
            df.drop(columns=['destinatario_norm', 'destinatario_homologado'], inplace=True)

            # === FIN HOMOLOGACIÓN =================================================================
            #-----------

            agrupado = df.groupby(['numero_externo', 'sku', 'bodega', 'nombre_bodega', 'destinatario','pdv'], as_index=False)['cantidad'].sum()
            bloques = empaquetar_ordenes_optimo(agrupado)

            for _, bloque in bloques:
                resumen_global.append(bloque)

            archivos = exportar_bloques_a_template(bloques, plantilla_path, carpeta_salida, hoja)
            archivos_exportados.extend(archivos)

        #resumen_final = pd.DataFrame()
        # Resumen
        if resumen_global:
            df_resumen = pd.concat(resumen_global, ignore_index=True)
            resumen_final = df_resumen.groupby('nombre_bodega').agg(
                skus_enviados=('sku', 'nunique'),
                ordenes_enviadas=('numero_externo', 'nunique'),
                cantidad_total=('cantidad', 'sum')
            ).reset_index()
            resumen_path = os.path.join(carpeta_salida, 'resumen_final.xlsx')
            resumen_final.to_excel(resumen_path, index=False)
            archivos_exportados.append(resumen_path)

        # Copiar ZIP fuera del with
        zip_final = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
        with zipfile.ZipFile(zip_final.name, 'w') as zipf:
            for archivo in archivos_exportados:
                zipf.write(archivo, os.path.basename(archivo))

        return zip_final.name, resumen_final

# ==== Streamlit UI ====
st.set_page_config(page_title="Generador de archivos Medipiel", layout="wide")
st.title("🧾 Generador de archivos Medipiel")

archivo_medipiel = st.file_uploader("📥 Sube el archivo de Medipiel (.xlsx):", type="xlsx")
plantilla_path = "template.xlsx"  # ✅ Cargada desde tu directorio

if archivo_medipiel:
    if st.button("🚀 Procesar archivo"):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_medipiel:
            tmp_medipiel.write(archivo_medipiel.read())
            medipiel_path = tmp_medipiel.name

        zip_resultado, resumen_df = procesar_archivo_medipiel(medipiel_path, plantilla_path)

        with open(zip_resultado, "rb") as f:
            st.download_button("📦 Descargar resultados (.zip)", f, file_name="resultado_medipiel.zip")

        st.success("✅ Archivos generados exitosamente")
        st.subheader("📊 Resumen por bodega")
        st.dataframe(resumen_df)
else:
    st.info("Por favor, sube el archivo de Medipiel para continuar.")
